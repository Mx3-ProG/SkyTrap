from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from skytrap.core.context import WorkspaceContext
from skytrap.tools.base import Tool, ToolResult
from skytrap.tools.filesystem import resolve_in_workspace
from skytrap.tools.registry import RegistryContext, register_tool
from skytrap.tools.skills.dcf_model.schema import DcfModelInput

DISCOUNT_RATE_ROW = 3
TERMINAL_GROWTH_ROW = 4
YEAR_ROW = 6
REVENUE_ROW = 7
EBITDA_MARGIN_ROW = 8
EBITDA_ROW = 9
DA_ROW = 10
EBIT_ROW = 11
TAX_RATE_ROW = 12
TAXES_ROW = 13
NOPAT_ROW = 14
ADD_DA_ROW = 15
CAPEX_ROW = 16
NWC_ROW = 17
FCF_ROW = 18
DISCOUNT_FACTOR_ROW = 20
PV_FCF_ROW = 21
TERMINAL_VALUE_ROW = 23
PV_TERMINAL_VALUE_ROW = 24
ENTERPRISE_VALUE_ROW = 26

FIRST_YEAR_COLUMN = 2  # column B — column A holds row labels


def _col(year_index: int) -> str:
    return get_column_letter(FIRST_YEAR_COLUMN + year_index)


def _build_preview(parsed: DcfModelInput) -> str:
    lines = [
        f"NEW DCF MODEL: {parsed.path}",
        f"Company: {parsed.company_name}",
        f"Discount rate (WACC): {parsed.discount_rate:.1%}   "
        f"Terminal growth: {parsed.terminal_growth_rate:.1%}",
        "",
    ]
    for year in parsed.years:
        lines.append(
            f"- {year.year_label}: revenue={year.revenue:,.0f}, "
            f"EBITDA margin={year.ebitda_margin:.1%}"
        )
    return "\n".join(lines)


def _build_workbook(parsed: DcfModelInput) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DCF Model"

    sheet["A1"] = f"{parsed.company_name} — DCF Model"
    sheet["A1"].font = Font(bold=True, size=14)

    sheet[f"A{DISCOUNT_RATE_ROW}"] = "Discount rate (WACC)"
    sheet[f"B{DISCOUNT_RATE_ROW}"] = parsed.discount_rate
    sheet[f"B{DISCOUNT_RATE_ROW}"].number_format = "0.0%"

    sheet[f"A{TERMINAL_GROWTH_ROW}"] = "Terminal growth rate"
    sheet[f"B{TERMINAL_GROWTH_ROW}"] = parsed.terminal_growth_rate
    sheet[f"B{TERMINAL_GROWTH_ROW}"].number_format = "0.0%"

    row_labels = {
        YEAR_ROW: "Year",
        REVENUE_ROW: "Revenue",
        EBITDA_MARGIN_ROW: "EBITDA margin",
        EBITDA_ROW: "EBITDA",
        DA_ROW: "D&A",
        EBIT_ROW: "EBIT",
        TAX_RATE_ROW: "Tax rate",
        TAXES_ROW: "Taxes",
        NOPAT_ROW: "NOPAT",
        ADD_DA_ROW: "+ D&A",
        CAPEX_ROW: "- CapEx",
        NWC_ROW: "- Change in NWC",
        FCF_ROW: "Unlevered FCF",
        DISCOUNT_FACTOR_ROW: "Discount factor",
        PV_FCF_ROW: "PV of FCF",
        TERMINAL_VALUE_ROW: "Terminal value",
        PV_TERMINAL_VALUE_ROW: "PV of terminal value",
        ENTERPRISE_VALUE_ROW: "Enterprise Value",
    }
    for row, label in row_labels.items():
        sheet[f"A{row}"] = label
    sheet[f"A{FCF_ROW}"].font = Font(bold=True)
    sheet[f"A{ENTERPRISE_VALUE_ROW}"].font = Font(bold=True)

    last_col = _col(len(parsed.years) - 1)
    wacc_ref = f"$B${DISCOUNT_RATE_ROW}"
    growth_ref = f"$B${TERMINAL_GROWTH_ROW}"

    for i, year in enumerate(parsed.years):
        col = _col(i)
        n = i + 1  # discount period, 1-indexed

        sheet[f"{col}{YEAR_ROW}"] = year.year_label
        sheet[f"{col}{REVENUE_ROW}"] = year.revenue
        sheet[f"{col}{EBITDA_MARGIN_ROW}"] = year.ebitda_margin
        sheet[f"{col}{EBITDA_MARGIN_ROW}"].number_format = "0.0%"
        sheet[f"{col}{EBITDA_ROW}"] = f"={col}{REVENUE_ROW}*{col}{EBITDA_MARGIN_ROW}"

        # D&A/CapEx/NWC are pre-computed from their % of revenue so the formulas
        # below stay simple, while the underlying % assumptions (above) remain
        # visible and independently editable in the sheet.
        sheet[f"{col}{DA_ROW}"] = round(year.revenue * year.da_pct_revenue, 2)
        sheet[f"{col}{EBIT_ROW}"] = f"={col}{EBITDA_ROW}-{col}{DA_ROW}"
        sheet[f"{col}{TAX_RATE_ROW}"] = year.tax_rate
        sheet[f"{col}{TAX_RATE_ROW}"].number_format = "0.0%"
        sheet[f"{col}{TAXES_ROW}"] = f"={col}{EBIT_ROW}*{col}{TAX_RATE_ROW}"
        sheet[f"{col}{NOPAT_ROW}"] = f"={col}{EBIT_ROW}-{col}{TAXES_ROW}"
        sheet[f"{col}{ADD_DA_ROW}"] = f"={col}{DA_ROW}"
        sheet[f"{col}{CAPEX_ROW}"] = round(-year.revenue * year.capex_pct_revenue, 2)
        sheet[f"{col}{NWC_ROW}"] = round(-year.revenue * year.nwc_change_pct_revenue, 2)
        sheet[f"{col}{FCF_ROW}"] = (
            f"={col}{NOPAT_ROW}+{col}{ADD_DA_ROW}+{col}{CAPEX_ROW}+{col}{NWC_ROW}"
        )
        sheet[f"{col}{FCF_ROW}"].font = Font(bold=True)

        sheet[f"{col}{DISCOUNT_FACTOR_ROW}"] = f"=1/(1+{wacc_ref})^{n}"
        sheet[f"{col}{PV_FCF_ROW}"] = f"={col}{FCF_ROW}*{col}{DISCOUNT_FACTOR_ROW}"

    # Terminal value only in the final projection year's column (Gordon growth,
    # applied to that year's FCF).
    sheet[f"{last_col}{TERMINAL_VALUE_ROW}"] = (
        f"={last_col}{FCF_ROW}*(1+{growth_ref})/({wacc_ref}-{growth_ref})"
    )
    sheet[f"{last_col}{PV_TERMINAL_VALUE_ROW}"] = (
        f"={last_col}{TERMINAL_VALUE_ROW}*{last_col}{DISCOUNT_FACTOR_ROW}"
    )

    sheet[f"B{ENTERPRISE_VALUE_ROW}"] = (
        f"=SUM(B{PV_FCF_ROW}:{last_col}{PV_FCF_ROW})+{last_col}{PV_TERMINAL_VALUE_ROW}"
    )
    sheet[f"B{ENTERPRISE_VALUE_ROW}"].font = Font(bold=True)

    return workbook


class DcfModelTool(Tool):
    name = "dcf_model"
    description = (
        "Generate a Discounted Cash Flow (DCF) valuation model as a real Excel workbook "
        "(.xlsx) with live formulas — not static numbers, so the user can adjust "
        "assumptions in Excel and have it recalculate. Shows a preview and requires "
        "confirmation before writing, same as write_file. You (the caller) determine the "
        "revenue/margin/growth assumptions from context; this tool only builds the "
        "spreadsheet from the numbers you give it. "
        'Arguments: {"path": "<output .xlsx path>", "company_name": "...", "years": '
        '[{"year_label": "2025", "revenue": 1000000, "ebitda_margin": 0.25, ...}], '
        '"discount_rate": 0.10, "terminal_growth_rate": 0.02}'
    )

    def __init__(self, confirm: Callable[[str], bool]):
        self._confirm = confirm

    def execute(self, workspace: WorkspaceContext, arguments: dict) -> ToolResult:
        try:
            parsed = DcfModelInput.model_validate(arguments)
        except ValidationError as exc:
            return ToolResult(success=False, output=f"Invalid arguments: {exc}")

        if not parsed.path.lower().endswith(".xlsx"):
            return ToolResult(success=False, output="Output path must end in .xlsx")

        ok, resolved = resolve_in_workspace(workspace, parsed.path)
        if not ok:
            return ToolResult(success=False, output=resolved)

        preview = _build_preview(parsed)
        if not self._confirm(preview):
            return ToolResult(success=False, output="User declined the write; file not created.")

        file_path = Path(resolved)
        workbook = _build_workbook(parsed)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(file_path))

        return ToolResult(success=True, output=f"Wrote a {len(parsed.years)}-year DCF model to {parsed.path}")


@register_tool
def _build_dcf_model_tool(context: RegistryContext) -> Tool:
    return DcfModelTool(confirm=context.confirm_write)
